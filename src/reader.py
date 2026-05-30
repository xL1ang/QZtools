"""读取 input/ 子文件夹中的所有文档，统一为文本。"""

from pathlib import Path

import docx
import fitz  # pymupdf
import openpyxl

INPUT_DIR = Path(__file__).parent.parent / "input"

# 支持的文本格式
TEXT_EXTENSIONS = {".md", ".txt", ".html", ".htm"}
# 首版跳过的格式
SKIP_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp"}


def read_text_file(path: Path) -> str:
    """读取纯文本文件。"""
    return path.read_text(encoding="utf-8")


def read_docx(path: Path) -> str:
    """读取 .docx 文件，提取所有段落文本。"""
    doc = docx.Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def read_pdf(path: Path) -> str:
    """读取 .pdf 文件，提取所有页面文本。"""
    text_parts = []
    with fitz.open(str(path)) as pdf:
        for page in pdf:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


def read_xlsx(path: Path) -> str:
    """读取 .xlsx 文件，提取所有工作表内容。"""
    wb = openpyxl.load_workbook(str(path), read_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text_parts.append(f"[工作表: {sheet_name}]")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            text_parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(text_parts)


def read_file(path: Path) -> str | None:
    """根据文件扩展名选择解析方式，返回文本内容。"""
    suffix = path.suffix.lower()

    if suffix in TEXT_EXTENSIONS:
        return read_text_file(path)
    elif suffix == ".docx":
        return read_docx(path)
    elif suffix == ".pdf":
        return read_pdf(path)
    elif suffix == ".xlsx":
        return read_xlsx(path)
    elif suffix in SKIP_EXTENSIONS:
        return None
    else:
        # 未知格式尝试作为文本读取
        try:
            return read_text_file(path)
        except (UnicodeDecodeError, Exception):
            return None


def read_folder(folder: Path) -> str:
    """读取一个文件夹下所有文件，拼成带文件名标记的文本。"""
    if not folder.exists():
        return ""

    parts = []
    for file_path in sorted(folder.iterdir()):
        if file_path.is_file() and file_path.name != ".gitkeep":
            try:
                content = read_file(file_path)
                if content:
                    parts.append(f"--- {file_path.name} ---\n{content}")
            except Exception as e:
                print(f"  [..] 跳过无法解析的文件: {file_path.name} ({e})")

    return "\n\n".join(parts)


def read_all_inputs() -> dict[str, str]:
    """读取 input/ 下四个子文件夹的所有文档，返回 dict。"""
    return {
        "existing_logic": read_folder(INPUT_DIR / "existing-logic"),
        "interfaces": read_folder(INPUT_DIR / "interfaces"),
        "prototypes": read_folder(INPUT_DIR / "prototypes"),
        "format_standards": read_folder(INPUT_DIR / "format-standards"),
    }
